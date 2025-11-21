import streamlit as st
import openpyxl
import re
import unicodedata
from io import BytesIO
import pandas as pd

# 1) Page config
st.set_page_config(page_title="Search App", page_icon="🔎", layout="centered")

if 'authenticated' not in st.session_state or not st.session_state.authenticated:
    st.switch_page("app.py")

# ========== 1) CSS base ==========
st.markdown("""
<style>
/* Nascondi il menu interno della sidebar */
[data-testid="stSidebarNav"] {
    display: none !important;
}

/* Stile generale sidebar */
[data-testid="stSidebar"] {
    width: 550px !important;
    min-width: 550px !important;
    max-width: 550px !important;
    background-color: #ecf0f1 !important;
    padding: 10px !important;
    transition: transform 0.4s ease-in-out !important;
    box-shadow: 2px 0 8px rgba(0, 0, 0, 0.2);
    z-index: 1000 !important;
}

/* Sfondo principale */
section.main {
    background-color: #d8dfe6 !important;
}
.main .block-container, 
div[data-testid="stAppViewContainer"] > section > div.block-container {
    background-color: transparent !important;
    padding: 2rem 1rem 1rem 1rem !important;
}

/* Nasconde completamente la sidebar */
.sidebar-hidden {
    transform: translateX(-100%) !important;
}

/* --- Stile pulsante freccia --- */
div[data-testid="collapsedControl"] {
    display: flex !important;
    align-items: center !important;
    gap: 6px !important;
}
div[data-testid="collapsedControl"] button {
    background-color: #f39c12 !important;
    border-radius: 50% !important;
    box-shadow: 0 0 6px rgba(0,0,0,0.3);
    transition: all 0.3s ease-in-out !important;
}
div[data-testid="collapsedControl"] button:hover {
    background-color: #e67e22 !important;
    transform: scale(1.15);
}

/* --- Testo accanto alla freccia --- */
#sidebar-toggle-label {
    color: #34495e;
    font-weight: 600;
    font-family: "Segoe UI", sans-serif;
    font-size: 14px;
    user-select: none;
}
</style>
""", unsafe_allow_html=True)


# ========== 2) Sidebar ==========
st.sidebar.page_link("app.py", label="**PDM Utility Hub**", icon="🏠")
st.sidebar.markdown("---")


# ========== 3) JavaScript per chiusura e testo dinamico ==========
st.markdown("""
<script>
const observer = new MutationObserver(() => {
  const container = window.parent.document.querySelector('div[data-testid="collapsedControl"]');
  const sidebar = window.parent.document.querySelector('[data-testid="stSidebar"]');
  if (container && sidebar) {
    observer.disconnect();

    // Aggiunge testo accanto alla freccia se non esiste
    if (!container.querySelector('#sidebar-toggle-label')) {
      const label = document.createElement('span');
      label.id = 'sidebar-toggle-label';
      label.innerText = 'Click to collapse';
      container.appendChild(label);
    }

    const btn = container.querySelector('button');
    const label = container.querySelector('#sidebar-toggle-label');

    btn.addEventListener('click', () => {
      if (sidebar.classList.contains('sidebar-hidden')) {
        sidebar.classList.remove('sidebar-hidden');
        label.innerText = 'Click to collapse';
      } else {
        sidebar.classList.add('sidebar-hidden');
        label.innerText = 'Click to expand';
      }
    });
  }
});

// Osserva eventuali cambiamenti nel DOM finché non trova gli elementi
observer.observe(window.parent.document.body, { childList: true, subtree: true });
</script>
""", unsafe_allow_html=True)

# ------------ Helpers & State ------------

def strip_accents(s):
    if not isinstance(s, str):
        s = str(s)
    return ''.join(ch for ch in unicodedata.normalize('NFD', s) if not unicodedata.combining(ch))

def build_spacing_pattern(term):
    return r"(?<!\\w)" + ''.join([re.escape(c) + r"\\s*" for c in term]) + r"(?!\\w)"

# Inizializza lo stato per input e uploader
if 'uploader_key' not in st.session_state:
    st.session_state['uploader_key'] = 0

for i in range(1, 11):
    st.session_state.setdefault(f'term{i}', '')

st.session_state.setdefault('custom_filename', 'filtered_results')
st.session_state.setdefault('download_bytes', b'')
st.session_state.setdefault('download_filename', '')

def clear_all():
    # Svuota TUTTI i campi delle celle
    for i in range(1, 11):
        st.session_state[f'term{i}'] = ''
    # Svuota filename e download
    st.session_state['custom_filename'] = ''
    st.session_state['download_bytes'] = b''
    st.session_state['download_filename'] = ''
    # Resetta anche il file uploader
    st.session_state['uploader_key'] += 1

# ------------ UI ------------
st.title("Search App:")

st.markdown("""
**ℹ️ How to use:**
- Upload an Excel file (.xlsx or .xls).
- Enter up to **10** search terms or phrases.
- Choose the final output filename.
- Click **Search and Download** to get the filtered Excel + report.
""")

# Clear: svuota tutto e ricarica la pagina
if st.button("Clear cache and data"):
    clear_all()
    st.rerun()

# Uploader con chiave variabile per poterlo resettare
uploaded_file = st.file_uploader(
    "Choose an Excel file",
    type=["xlsx", "xls"],
    key=f'uploaded_file_{st.session_state["uploader_key"]}'
)

# Inputs vincolati allo stato
for i in range(1, 11):
    st.text_input(f"Term {i}", key=f'term{i}')

custom_filename = st.text_input("Output filename", key="custom_filename")

# ------------ Azione: cerca e prepara il file ------------
if st.button("Search and Download"):
    # Costruisci lista termini puliti
    terms = [st.session_state[f'term{i}'].strip() for i in range(1, 11) if st.session_state[f'term{i}'].strip()]

    if not uploaded_file:
        st.error("Please upload a file first.")
        st.stop()

    if not terms:
        st.error("Please enter at least one search term.")
        st.stop()

    st.info("Reading file progressively — please wait...")

    # compile patterns
    term_noacc = [strip_accents(t) for t in terms]
    term_compact = [re.sub(r"\\s+", "", t) for t in term_noacc]
    compiled = [re.compile(build_spacing_pattern(t), re.IGNORECASE) for t in term_compact]
    per_term_counts = [0] * len(compiled)

    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    matches = []
    total = 0
    matched = 0
    progress = st.progress(0)

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total += 1
        row_values = ["" if v is None else str(v) for v in row]
        text = ' '.join(row_values)
        text_noacc = strip_accents(text)
        row_hits = [terms[j] for j, pat in enumerate(compiled) if pat.search(text_noacc)]

        if row_hits:
            matched += 1
            for j, pat in enumerate(compiled):
                if pat.search(text_noacc):
                    per_term_counts[j] += 1

            row_dict = dict(zip(headers, row_values))
            row_dict["Matched terms"] = ";".join(row_hits)
            matches.append(row_dict)

        if i % 1000 == 0:
            progress.progress(min(1.0, i / (ws.max_row or 1)))

    progress.empty()
    wb.close()

    if not matches:
        st.warning("No matches found.")
        st.stop()

    result_df = pd.DataFrame(matches)
    report_df = pd.DataFrame({"Term": terms, "Rows matched": per_term_counts})

    st.success(f"Matched {matched} rows out of ~{total} scanned.")
    st.dataframe(report_df, use_container_width=True)

    # Buffer -> salva in sessione per tenere il bottone
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="Filtered", index=False)
        report_df.to_excel(writer, sheet_name="Report", index=False)
    buf.seek(0)

    safe_name = re.sub(r'[<>:"/\\\\|?*]', '_', (st.session_state['custom_filename'] or '').strip()) or "filtered_results"
    st.session_state['download_bytes'] = buf.getvalue()
    st.session_state['download_filename'] = f"{safe_name}.xlsx"

# ------------ Download persistente ------------
if st.session_state.get('download_bytes'):
    st.download_button(
        "Download filtered Excel + report",
        data=st.session_state['download_bytes'],
        file_name=st.session_state['download_filename'] or "filtered_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_btn",
        use_container_width=True
    )
